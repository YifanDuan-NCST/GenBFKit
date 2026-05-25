from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List

import pandas as pd

from .dictionary.base_dictionary import BaseDictionary
from .dictionary.data_attribute_dictionary import DataAttributeDictionary
from .dictionary.data_category_dictionary import DataCategoryDictionary
from .dictionary.data_pool_dictionary import DataPoolDictionary
from .dictionary.dataset_dictionary import DatasetDictionary
from .dictionary.registry import normalize_text
from . import prebuilt_default


class GenBFKitDictManager:
    """
    A small orchestration layer over the 5 independent dictionaries.

    Notes:
    - This folder is named `Construction_data_dictionary`. Older examples may reference a
      non-existent `genbfkit_dictionary` package; use this manager instead.
    - This repository snapshot doesn't include the Excel sources, so `load_from_project_root`
      is best-effort and will return zeros if no source files are found.
    """

    def __init__(self) -> None:
        self.base_dict = BaseDictionary()
        self.category_dict = DataCategoryDictionary()
        self.pool_dict = DataPoolDictionary()
        self.dataset_dict = DatasetDictionary()
        self.attr_dict = DataAttributeDictionary()

    def load_from_project_root(self, project_root: str | Path, *, overwrite: bool = True) -> Dict[str, int]:
        """
        Best-effort loader for typical Excel dictionary files if present.

        Returns a dict of loaded row counts per dictionary.
        """
        project_root = Path(project_root)

        # Common filenames (kept flexible; we try multiple)
        candidates = {
            "base_dictionary": ["base_dictionary.xlsx", "base_dictionary - 示例.xlsx"],
            "data_category_dictionary": ["data_category_dictionary.xlsx"],
            "data_pool_dictionary": ["data_pool_dictionary.xlsx"],
            "dataset_dictionary": ["dataset_dictionary.xlsx", "dataset_dictionary - 示例.xlsx"],
            "data_attribute_dictionary": ["data_attribute_dictionary.xlsx"],
        }

        def _find_any(names: List[str]) -> Path | None:
            for n in names:
                p = project_root / n
                if p.exists():
                    return p
                p2 = project_root / "Construction_data_dictionary" / n
                if p2.exists():
                    return p2
            return None

        counts: Dict[str, int] = {k: 0 for k in candidates}
        found_any = False

        p = _find_any(candidates["base_dictionary"])
        if p:
            found_any = True
            counts["base_dictionary"] = self.base_dict.load_from_excel(p, overwrite=overwrite)

        p = _find_any(candidates["data_category_dictionary"])
        if p:
            found_any = True
            counts["data_category_dictionary"] = self.category_dict.load_from_excel(p, overwrite=overwrite)

        p = _find_any(candidates["data_pool_dictionary"])
        if p:
            found_any = True
            counts["data_pool_dictionary"] = self.pool_dict.load_from_excel(p, overwrite=overwrite)

        p = _find_any(candidates["dataset_dictionary"])
        if p:
            found_any = True
            counts["dataset_dictionary"] = self.dataset_dict.load_from_excel(p, overwrite=overwrite)

        p = _find_any(candidates["data_attribute_dictionary"])
        if p:
            found_any = True
            counts["data_attribute_dictionary"] = self.attr_dict.load_from_excel(p, overwrite=overwrite)

        # If nothing is found, fall back to built-in prebuilt skeleton data
        if not found_any:
            counts["prebuilt_default"] = self.load_prebuilt_default(overwrite=overwrite)

        return counts

    def load_prebuilt_default(self, *, overwrite: bool = True) -> int:
        """
        Load the built-in minimal prebuilt dictionary data shipped with the code.

        Returns number of inserted items across dictionaries (best-effort count).
        """
        n = 0

        for w in prebuilt_default.base_work_types():
            self.base_dict.add(w["work_type_en"], work_type_zh=w.get("work_type_zh", ""), no=w.get("no"), overwrite=overwrite)
            n += 1

        for c in prebuilt_default.category_items():
            self.category_dict.add(c["work_type_en"], c["category_en"], category_zh=c.get("category_zh", ""), overwrite=overwrite)
            n += 1

        for p in prebuilt_default.pool_items():
            # Global pool types (9 categories), not expanded by (work_type, category)
            self.pool_dict.add_pool_type(p["pool_en"], pool_zh=p.get("pool_zh", ""), overwrite=overwrite)
            n += 1

        for d in prebuilt_default.dataset_items():
            self.dataset_dict.add(
                d["work_type_en"],
                d["category_en"],
                d["pool_en"],
                d["dataset_en"],
                dataset_zh=d.get("dataset_zh", ""),
                dataset_zh_short=d.get("dataset_zh_short", ""),
                overwrite=overwrite,
            )
            n += 1

        for pool_type, attrs in prebuilt_default.attribute_templates():
            # ensure pool type exists, then set attributes
            self.attr_dict.add_pool_type(pool_type, overwrite=overwrite)
            for attr_id, attr_name in attrs.items():
                self.attr_dict.set_attribute(pool_type, attr_id, attr_name)
            n += 1

        return n

    def import_additional_data(self, excel_path: str | Path, *, overwrite: bool = False) -> Dict[str, int]:
        """
        Import additional data from Excel file and merge with existing data.

        This method supports TWO import modes:
        1. Single-file mode: Read from sheets within the Excel file (if sheets exist)
        2. Multi-file mode: Look for separate dictionary files in the same directory

        Priority: Single-file mode (sheets) > Multi-file mode (separate files)

        The key difference from load_from_project_root:
        - load_from_project_root: completely replaces existing data (overwrite=True by default)
        - import_additional_data: merges with existing data (overwrite=False by default)

        Args:
            excel_path: Path to the Excel file containing additional data
                      Can be either:
                      - A multi-sheet Excel file with sheets: Base Dict, Data Category, etc.
                      - A single Excel file that will trigger searching for separate dict files
            overwrite: If False, new items will be added without replacing existing ones.
                      If True, existing items with the same key will be overwritten.

        Returns:
            Dict with counts of imported items per dictionary type
        """
        import openpyxl

        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        counts: Dict[str, int] = {
            "base_dictionary": 0,
            "data_category_dictionary": 0,
            "data_pool_dictionary": 0,
            "dataset_dictionary": 0,
            "data_attribute_dictionary": 0,
        }

        # Sheet name mappings (English patterns to look for)
        sheet_mappings = {
            "base_dictionary": ["base", "work_type"],
            "data_category_dictionary": ["category", "data category"],
            "data_pool_dictionary": ["pool", "data pool"],
            "dataset_dictionary": ["dataset", "dataset_dict"],
            "data_attribute_dictionary": ["attribute", "attr", "attribute_dict"],
        }

        # Check if the Excel file contains sheets that match our dictionary types
        try:
            wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()

            # Try to find matching sheets in the Excel file
            matched_sheets = {}
            for dict_type, patterns in sheet_mappings.items():
                for sheet_name in sheet_names:
                    sheet_lower = sheet_name.lower()
                    for pattern in patterns:
                        if pattern.lower() in sheet_lower:
                            matched_sheets[dict_type] = sheet_name
                            break
                    if dict_type in matched_sheets:
                        break

            # If we found matching sheets, import from them
            if matched_sheets:
                if "base_dictionary" in matched_sheets:
                    counts["base_dictionary"] = self.base_dict.load_from_excel(
                        excel_path, sheet_name=matched_sheets["base_dictionary"], overwrite=overwrite
                    )

                if "data_category_dictionary" in matched_sheets:
                    counts["data_category_dictionary"] = self.category_dict.load_from_excel(
                        excel_path, sheet_name=matched_sheets["data_category_dictionary"], overwrite=overwrite
                    )

                if "data_pool_dictionary" in matched_sheets:
                    counts["data_pool_dictionary"] = self.pool_dict.load_from_excel(
                        excel_path, sheet_name=matched_sheets["data_pool_dictionary"], overwrite=overwrite
                    )

                if "dataset_dictionary" in matched_sheets:
                    counts["dataset_dictionary"] = self.dataset_dict.load_from_excel(
                        excel_path, sheet_name=matched_sheets["dataset_dictionary"], overwrite=overwrite
                    )

                if "data_attribute_dictionary" in matched_sheets:
                    counts["data_attribute_dictionary"] = self.attr_dict.load_from_excel(
                        excel_path, sheet_name=matched_sheets["data_attribute_dictionary"], overwrite=overwrite
                    )

                return counts

        except Exception:
            # If single-file mode fails, fall back to multi-file mode
            pass

        # Multi-file mode: Look for separate dictionary files
        _search_dirs = [excel_path.parent]
        if excel_path.is_file():
            _search_dirs.append(excel_path.parent.parent)
        else:
            _search_dirs = [excel_path]

        base_file = None
        for _dir in _search_dirs:
            base_file = self._find_file(_dir, ["base_dictionary.xlsx", "base_dictionary - 示例.xlsx"])
            if base_file:
                break
        if base_file:
            counts["base_dictionary"] = self.base_dict.load_from_excel(base_file, overwrite=overwrite)

        category_file = None
        for _dir in _search_dirs:
            category_file = self._find_file(_dir, ["data_category_dictionary.xlsx"])
            if category_file:
                break
        if category_file:
            counts["data_category_dictionary"] = self.category_dict.load_from_excel(category_file, overwrite=overwrite)

        pool_file = None
        for _dir in _search_dirs:
            pool_file = self._find_file(_dir, ["data_pool_dictionary.xlsx"])
            if pool_file:
                break
        if pool_file:
            counts["data_pool_dictionary"] = self.pool_dict.load_from_excel(pool_file, overwrite=overwrite)

        dataset_file = None
        for _dir in _search_dirs:
            dataset_file = self._find_file(_dir, ["dataset_dictionary.xlsx", "dataset_dictionary - 示例.xlsx"])
            if dataset_file:
                break
        if dataset_file:
            counts["dataset_dictionary"] = self.dataset_dict.load_from_excel(dataset_file, overwrite=overwrite)

        attr_file = None
        for _dir in _search_dirs:
            attr_file = self._find_file(_dir, ["data_attribute_dictionary.xlsx"])
            if attr_file:
                break
        if attr_file:
            counts["data_attribute_dictionary"] = self.attr_dict.load_from_excel(attr_file, overwrite=overwrite)

        return counts

    def import_from_simple_template(
        self,
        excel_path: str | Path,
        *,
        work_type_zh: str = "",
        category_zh: str = "",
        overwrite: bool = False,
    ) -> Dict[str, int]:
        """
        Import additional data from a simple 4-column template.

        This method is designed for the simple template format:
        - Work type
        - data category
        - data pool
        - dataset

        Each row represents a new dataset entry. The data pool must be one of the
        existing 9 pool types (which are loaded from prebuilt_default).

        Usage:
            # 1. Load prebuilt data first
            mgr.load_prebuilt_default()

            # 2. Import additional data from simple template
            counts = mgr.import_from_simple_template("my_custom_data.xlsx")

        Args:
            excel_path: Path to the Excel file with 4 columns:
                       [Work type, data category, data pool, dataset]
            work_type_zh: Default Chinese name for work types (optional)
            category_zh: Default Chinese name for categories (optional)
            overwrite: If False, skip existing keys. If True, overwrite existing data.

        Returns:
            Dict with counts: {
                "work_types_added": N,
                "categories_added": N,
                "datasets_added": N,
                "rows_skipped": N,
            }
        """
        import pandas as pd

        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Read Excel file - find the "Simple Template" sheet by index or name pattern
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_names = wb.sheetnames
        wb.close()

        # Find "Simple Template" sheet (usually the 6th sheet, index 5)
        simple_sheet_idx = None
        for i, name in enumerate(sheet_names):
            # Check if any byte pattern matches "Simple" or "4" in the sheet name
            name_lower = name.lower()
            if "simple" in name_lower or "4" in name_lower or "4列" in name:
                simple_sheet_idx = i
                break

        # Default to last sheet which is usually "Simple Template" for this file
        if simple_sheet_idx is None:
            simple_sheet_idx = len(sheet_names) - 1

        # Excel structure: Row 1=title, Row 2=description, Row 3=headers, Row 4=descriptions, Row 5+=data
        # With header=2, pandas uses row 3 (0-indexed row 2) as column names
        # skiprows=[3] skips row 4 (description row)
        df = pd.read_excel(excel_path, sheet_name=simple_sheet_idx, header=2, skiprows=[3])

        # Clean up column names - strip whitespace and normalize
        df.columns = [str(c).strip() for c in df.columns]

        # Expected columns (case-insensitive matching)
        # Template may have extra columns like "No.", "是否预构建", "备注"
        col_wt = next((c for c in df.columns if c.lower().strip() in {"work type", "work_type", "worktype"}), None)
        col_cat = next((c for c in df.columns if "data category" in c.lower().strip() or c.lower().strip() == "data_category"), None)
        col_pool = next((c for c in df.columns if "data pool" in c.lower().strip() or c.lower().strip() == "data_pool"), None)
        col_dataset = next((c for c in df.columns if c.lower().strip() == "dataset"), None)

        # If we found the 4 key columns, we can proceed
        if col_wt is None or col_cat is None or col_pool is None or col_dataset is None:
            raise ValueError(
                f"Template must have 4 columns: Work Type, Data Category, Data Pool, Dataset. "
                f"Found columns: {list(df.columns)}"
            )

        # Counters
        work_types_added = 0
        categories_added = 0
        datasets_added = 0
        rows_skipped = 0

        # Track existing entries to avoid duplicates
        existing_work_types = set(self.get_work_types())
        existing_categories = {(c.work_type_en, c.category_en) for c in self.category_dict.list_all()}
        existing_datasets = {
            (d.work_type_en, d.category_en, d.pool_en, d.dataset_en)
            for d in self.dataset_dict.list_all()
        }

        # Get valid pool types from prebuilt data
        # Template uses underscores (e.g., Continuous_time_series_data)
        # Prebuilt data uses spaces (e.g., Continuous time-series data)
        valid_pool_names = set(self.pool_dict.list_pool_types())
        pool_name_mapping = {}
        for pool in valid_pool_names:
            # Normalize: replace spaces with underscore, replace hyphens with underscore
            underscore_version = pool.replace(" ", "_").replace("-", "_")
            pool_name_mapping[underscore_version] = pool
            pool_name_mapping[pool] = pool  # Also keep original

        # Calculate max work type number for auto-increment
        all_work_types = list(self.base_dict.list_all())
        max_no = 0
        for wt in all_work_types:
            if wt.no is not None and wt.no > max_no:
                max_no = wt.no

        # Process each row
        for _, row in df.iterrows():
            work_type = normalize_text(row.get(col_wt))
            category = normalize_text(row.get(col_cat))
            pool_raw = normalize_text(row.get(col_pool))
            dataset = normalize_text(row.get(col_dataset))

            # Skip empty rows (including description rows that have no valid data)
            if not all([work_type, category, pool_raw, dataset]):
                rows_skipped += 1
                continue

            # Normalize pool type name (underscore to space format, also handle hyphens)
            pool_normalized = pool_raw.replace(" ", "_").replace("-", "_")
            pool = pool_name_mapping.get(pool_normalized, pool_name_mapping.get(pool_raw, pool_raw))
            if pool not in valid_pool_names:
                print(f"Warning: Invalid pool type '{pool_raw}' at row {_ + 1}. Must be one of: {valid_pool_names}")
                rows_skipped += 1
                continue

            # Add work type if not exists (with auto-increment number)
            if work_type not in existing_work_types:
                try:
                    max_no += 1  # Auto-increment number
                    self.base_dict.add(work_type, work_type_zh=work_type_zh, no=max_no, overwrite=overwrite)
                    existing_work_types.add(work_type)
                    work_types_added += 1
                except Exception:
                    pass  # Already exists or error

            # Add category if not exists
            category_key = (work_type, category)
            if category_key not in existing_categories:
                try:
                    self.category_dict.add(work_type, category, category_zh=category_zh, overwrite=overwrite)
                    existing_categories.add(category_key)
                    categories_added += 1
                except Exception:
                    pass  # Already exists or error

            # Add dataset
            dataset_key = (work_type, category, pool, dataset)
            if dataset_key not in existing_datasets or overwrite:
                try:
                    self.dataset_dict.add(
                        work_type,
                        category,
                        pool,
                        dataset,
                        dataset_zh="",
                        dataset_zh_short="",
                        overwrite=overwrite,
                    )
                    existing_datasets.add(dataset_key)
                    datasets_added += 1
                except Exception as e:
                    # Check if it's a duplicate key error
                    if "duplicate" in str(e).lower():
                        rows_skipped += 1
                    else:
                        print(f"Error adding dataset '{dataset}': {e}")
                        rows_skipped += 1
            else:
                rows_skipped += 1

        return {
            "work_types_added": work_types_added,
            "categories_added": categories_added,
            "datasets_added": datasets_added,
            "rows_skipped": rows_skipped,
        }

    def load_text_data_template(self, excel_path: str | Path) -> "pd.DataFrame":
        """
        Load Text data storage template from Excel file.

        This method reads the Text data template which contains:
        - dataset_id: Dataset ID
        - work_type: Work type
        - data_category: Data category
        - data_pool: Data pool type (Text data)
        - english_name: English name
        - chinese_name: Chinese name
        - data_description: Data description
        - data_storage_type: Storage type (TEXT/MEDIUMTEXT/LONGTEXT)
        - storage_location: Storage location
        - text_encoding: Text encoding (UTF-8/GBK/GB2312)
        - text_format: Text format (JSON/XML/CSV/Plain)
        - max_length: Maximum length
        - min_length: Minimum length
        - priority_level: Priority level (1-5)
        - creation_time: Creation time
        - keyword_set: Keyword set (comma-separated)
        - annotation_label: Annotation label (comma-separated)
        - text_content: Actual text content
        - remarks: Remarks
        - last_updated: Last updated time

        Args:
            excel_path: Path to the Text data template Excel file

        Returns:
            DataFrame containing the Text data template
        """
        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        df = pd.read_excel(excel_path, sheet_name="Text data storage")
        return df

    def load_image_data_template(self, excel_path: str | Path) -> "pd.DataFrame":
        """
        Load Image data storage template from Excel file.

        This method reads the Image data template which contains:
        - dataset_id: Dataset ID
        - work_type: Work type
        - data_category: Data category
        - data_pool: Data pool type (Image data)
        - english_name: English name
        - chinese_name: Chinese name
        - data_description: Data description
        - data_storage_type: Storage type (BLOB/MEDIUMBLOB/LONGBLOB/文件路径)
        - storage_location: Storage location
        - image_resolution: Image resolution (e.g., 1920x1080)
        - image_format: Image format (JPEG/PNG/BMP/TIFF)
        - color_mode: Color mode (RGB/灰度/CMYK)
        - blur_threshold: Blur threshold
        - file_size_limit: File size limit (MB)
        - priority_level: Priority level (1-5)
        - creation_time: Creation time
        - annotation_type: Annotation type (分类/检测/分割/关键点)
        - label_name: Label name
        - image_path: Image path (local path/OSS URL)
        - image_base64: Image Base64 encoding
        - remarks: Remarks
        - last_updated: Last updated time

        Args:
            excel_path: Path to the Image data template Excel file

        Returns:
            DataFrame containing the Image data template
        """
        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        df = pd.read_excel(excel_path, sheet_name="Image data storage")
        return df

    def import_text_data_template(self, excel_path: str | Path, *, overwrite: bool = False) -> Dict[str, int]:
        """
        Import Text data from Excel template and add to dictionary system.

        This method reads the Text data template and adds entries to:
        - base_dict: work types
        - category_dict: data categories
        - pool_dict: data pools (Text data)
        - dataset_dict: datasets

        Args:
            excel_path: Path to the Text data template Excel file
            overwrite: If False, skip existing keys. If True, overwrite existing data.

        Returns:
            Dict with counts: {
                "work_types_added": N,
                "categories_added": N,
                "pools_added": N,
                "datasets_added": N,
                "rows_skipped": N,
            }
        """
        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        df = pd.read_excel(excel_path, sheet_name="Text data storage")

        # Normalize column names
        df.columns = [normalize_text(c) for c in df.columns]

        # Counters
        work_types_added = 0
        categories_added = 0
        pools_added = 0
        datasets_added = 0
        rows_skipped = 0

        # Track existing entries to avoid duplicates
        existing_work_types = set(self.get_work_types())
        existing_categories = {(c.work_type_en, c.category_en) for c in self.category_dict.list_all()}
        existing_pools = {(p.work_type_en, p.category_en, p.pool_en) for p in self.pool_dict.list_all()}
        existing_datasets = {
            (d.work_type_en, d.category_en, d.pool_en, d.dataset_en)
            for d in self.dataset_dict.list_all()
        }

        # Ensure Text data pool type exists
        text_pool_type = "Text data"
        try:
            self.pool_dict.add_pool_type(text_pool_type, pool_zh="文本数据", overwrite=False)
            pools_added += 1
        except Exception:
            pass

        # Get column mappings
        col_wt = next((c for c in df.columns if c.lower() in {"work_type", "work type"}), None)
        col_cat = next((c for c in df.columns if "data_category" in c.lower() or "data category" in c.lower()), None)
        col_eng = next((c for c in df.columns if c.lower() in {"english_name", "english"}), None)
        col_zh = next((c for c in df.columns if c.lower() in {"chinese_name", "chinese"}), None)

        if col_wt is None or col_cat is None:
            raise ValueError(
                f"Text data template must have columns: work_type, data_category. "
                f"Found columns: {list(df.columns)}"
            )

        # Process each row
        for _, row in df.iterrows():
            work_type = normalize_text(row.get(col_wt))
            category = normalize_text(row.get(col_cat))
            english_name = normalize_text(row.get(col_eng)) if col_eng else ""
            chinese_name = normalize_text(row.get(col_zh)) if col_zh else ""

            # Skip empty rows
            if not work_type or not category:
                rows_skipped += 1
                continue

            # Add work type if not exists
            if work_type not in existing_work_types:
                try:
                    self.base_dict.add(work_type, work_type_zh="", overwrite=overwrite)
                    existing_work_types.add(work_type)
                    work_types_added += 1
                except Exception:
                    pass

            # Add category if not exists
            category_key = (work_type, category)
            if category_key not in existing_categories:
                try:
                    self.category_dict.add(work_type, category, category_zh="", overwrite=overwrite)
                    existing_categories.add(category_key)
                    categories_added += 1
                except Exception:
                    pass

            # Add pool entry (work_type -> category -> Text data pool)
            pool_key = (work_type, category, text_pool_type)
            if pool_key not in existing_pools:
                try:
                    self.pool_dict.add(work_type, category, text_pool_type, overwrite=overwrite)
                    existing_pools.add(pool_key)
                except Exception:
                    pass

            # Add dataset (english_name or dataset_id as dataset_en)
            if english_name:
                dataset_en = english_name
            else:
                dataset_id = normalize_text(row.get("dataset_id", ""))
                dataset_en = dataset_id if dataset_id else f"text_{rows_skipped}"

            dataset_key = (work_type, category, text_pool_type, dataset_en)
            if dataset_key not in existing_datasets or overwrite:
                try:
                    self.dataset_dict.add(
                        work_type,
                        category,
                        text_pool_type,
                        dataset_en,
                        dataset_zh=chinese_name,
                        dataset_zh_short=chinese_name,
                        overwrite=overwrite,
                    )
                    existing_datasets.add(dataset_key)
                    datasets_added += 1
                except Exception:
                    rows_skipped += 1
            else:
                rows_skipped += 1

        return {
            "work_types_added": work_types_added,
            "categories_added": categories_added,
            "pools_added": pools_added,
            "datasets_added": datasets_added,
            "rows_skipped": rows_skipped,
        }

    def import_image_data_template(self, excel_path: str | Path, *, overwrite: bool = False) -> Dict[str, int]:
        """
        Import Image data from Excel template and add to dictionary system.

        This method reads the Image data template and adds entries to:
        - base_dict: work types
        - category_dict: data categories
        - pool_dict: data pools (Image data)
        - dataset_dict: datasets

        Args:
            excel_path: Path to the Image data template Excel file
            overwrite: If False, skip existing keys. If True, overwrite existing data.

        Returns:
            Dict with counts: {
                "work_types_added": N,
                "categories_added": N,
                "pools_added": N,
                "datasets_added": N,
                "rows_skipped": N,
            }
        """
        from pathlib import Path as PathType

        excel_path = PathType(excel_path)

        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        df = pd.read_excel(excel_path, sheet_name="Image data storage")

        # Normalize column names
        df.columns = [normalize_text(c) for c in df.columns]

        # Counters
        work_types_added = 0
        categories_added = 0
        pools_added = 0
        datasets_added = 0
        rows_skipped = 0

        # Track existing entries to avoid duplicates
        existing_work_types = set(self.get_work_types())
        existing_categories = {(c.work_type_en, c.category_en) for c in self.category_dict.list_all()}
        existing_pools = {(p.work_type_en, p.category_en, p.pool_en) for p in self.pool_dict.list_all()}
        existing_datasets = {
            (d.work_type_en, d.category_en, d.pool_en, d.dataset_en)
            for d in self.dataset_dict.list_all()
        }

        # Ensure Image data pool type exists
        image_pool_type = "Image data"
        try:
            self.pool_dict.add_pool_type(image_pool_type, pool_zh="图像数据", overwrite=False)
            pools_added += 1
        except Exception:
            pass

        # Get column mappings
        col_wt = next((c for c in df.columns if c.lower() in {"work_type", "work type"}), None)
        col_cat = next((c for c in df.columns if "data_category" in c.lower() or "data category" in c.lower()), None)
        col_eng = next((c for c in df.columns if c.lower() in {"english_name", "english"}), None)
        col_zh = next((c for c in df.columns if c.lower() in {"chinese_name", "chinese"}), None)

        if col_wt is None or col_cat is None:
            raise ValueError(
                f"Image data template must have columns: work_type, data_category. "
                f"Found columns: {list(df.columns)}"
            )

        # Process each row
        for _, row in df.iterrows():
            work_type = normalize_text(row.get(col_wt))
            category = normalize_text(row.get(col_cat))
            english_name = normalize_text(row.get(col_eng)) if col_eng else ""
            chinese_name = normalize_text(row.get(col_zh)) if col_zh else ""

            # Skip empty rows
            if not work_type or not category:
                rows_skipped += 1
                continue

            # Add work type if not exists
            if work_type not in existing_work_types:
                try:
                    self.base_dict.add(work_type, work_type_zh="", overwrite=overwrite)
                    existing_work_types.add(work_type)
                    work_types_added += 1
                except Exception:
                    pass

            # Add category if not exists
            category_key = (work_type, category)
            if category_key not in existing_categories:
                try:
                    self.category_dict.add(work_type, category, category_zh="", overwrite=overwrite)
                    existing_categories.add(category_key)
                    categories_added += 1
                except Exception:
                    pass

            # Add pool entry (work_type -> category -> Image data pool)
            pool_key = (work_type, category, image_pool_type)
            if pool_key not in existing_pools:
                try:
                    self.pool_dict.add(work_type, category, image_pool_type, overwrite=overwrite)
                    existing_pools.add(pool_key)
                except Exception:
                    pass

            # Add dataset (english_name or dataset_id as dataset_en)
            if english_name:
                dataset_en = english_name
            else:
                dataset_id = normalize_text(row.get("dataset_id", ""))
                dataset_en = dataset_id if dataset_id else f"image_{rows_skipped}"

            dataset_key = (work_type, category, image_pool_type, dataset_en)
            if dataset_key not in existing_datasets or overwrite:
                try:
                    self.dataset_dict.add(
                        work_type,
                        category,
                        image_pool_type,
                        dataset_en,
                        dataset_zh=chinese_name,
                        dataset_zh_short=chinese_name,
                        overwrite=overwrite,
                    )
                    existing_datasets.add(dataset_key)
                    datasets_added += 1
                except Exception:
                    rows_skipped += 1
            else:
                rows_skipped += 1

        return {
            "work_types_added": work_types_added,
            "categories_added": categories_added,
            "pools_added": pools_added,
            "datasets_added": datasets_added,
            "rows_skipped": rows_skipped,
        }

    def get_pool_attributes(self, pool_type: str) -> Dict[str, str]:
        """
        Get attribute template for a specific pool type.

        Args:
            pool_type: Pool type (e.g., "Text data", "Image data")

        Returns:
            Dictionary of attribute_id -> attribute_name
        """
        try:
            return self.attr_dict.get_attributes_for_pool_type(pool_type)
        except Exception:
            return {}

    def _find_file(self, directory: Path, names: List[str]) -> Path | None:
        """Find a file from a list of possible names in the given directory."""
        for name in names:
            path = directory / name
            if path.exists():
                return path
        return None

    def export_to_json(self, output_path: str | Path, *, include_all: bool = True) -> None:
        """
        Export current dictionary data to JSON file.

        Args:
            output_path: Path to output JSON file
            include_all: If True, export all data. If False, export only non-prebuilt data.
        """
        from pathlib import Path as PathType

        output_path = PathType(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        export_data: Dict[str, Any] = {
            "base_work_types": [],
            "categories": [],
            "pools": [],
            "datasets": [],
            "attribute_templates": {},
        }

        # Export base_work_types
        for item in self.base_dict.list_all():
            export_data["base_work_types"].append({
                "work_type_en": item.work_type_en,
                "work_type_zh": item.work_type_zh or "",
                "no": item.no or 0,
            })

        # Export categories
        for item in self.category_dict.list_all():
            export_data["categories"].append({
                "work_type_en": item.work_type_en,
                "category_en": item.category_en,
                "category_zh": item.category_zh or "",
            })

        # Export pools — use list_pool_types() for the 9 unique global pool types
        seen_pools: set = set()
        for item in self.pool_dict.list_all():
            if item.pool_en and item.pool_en not in seen_pools:
                seen_pools.add(item.pool_en)
                export_data["pools"].append({
                    "pool_en": item.pool_en,
                    "pool_zh": item.pool_zh or "",
                })

        # Export datasets
        for item in self.dataset_dict.list_all():
            export_data["datasets"].append({
                "work_type_en": item.work_type_en,
                "category_en": item.category_en,
                "pool_en": item.pool_en,
                "dataset_en": item.dataset_en,
                "dataset_zh": item.dataset_zh or "",
                "dataset_zh_short": item.dataset_zh_short or "",
            })

        # Export attribute_templates
        for pool_type in self.attr_dict.list_pool_types():
            export_data["attribute_templates"][pool_type] = self.attr_dict.get_attributes_for_pool_type(pool_type)

        import json

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)

        print(f"Data exported to: {output_path}")

    def get_work_types(self) -> List[str]:
        """Return all known work types (English)."""
        return [w.work_type_en for w in self.base_dict.list_all()]

    def get_full_data_chain(self, work_type: str, *, include_attributes: bool = True) -> Dict[str, Any]:
        """
        Build a chain: work_type -> categories -> pools -> datasets.

        Attribute templates are only available by *pool_type* in this snapshot; unless the
        caller has a mapping from (pool/dataset) to pool_type, attributes will be empty.
        """
        wt = normalize_text(work_type)
        out: Dict[str, Any] = {"work_type": wt, "data_chain": []}

        for cat in self.category_dict.list_by_work_type(wt):
            cat_node: Dict[str, Any] = {
                "category_en": cat.category_en,
                "category_zh": cat.category_zh,
                "pools": [],
            }

            # Pools are global types; derive which pool types appear under this category from datasets.
            for pool_en in self.dataset_dict.list_pools_by_category(wt, cat.category_en):
                pool_item = self.pool_dict.try_get_pool_type(pool_en)
                pool_node: Dict[str, Any] = {
                    "pool_en": pool_en,
                    "pool_zh": pool_item.pool_zh if pool_item else "",
                    "datasets": [],
                }

                for ds in self.dataset_dict.list_by_pool(wt, cat.category_en, pool_en):
                    ds_node: Dict[str, Any] = {
                        "dataset_en": ds.dataset_en,
                        "dataset_zh": ds.dataset_zh,
                        "dataset_zh_short": ds.dataset_zh_short,
                    }
                    if include_attributes:
                        # In this repo snapshot, attribute templates are keyed by pool type.
                        # We treat `pool_en` as the pool type label (e.g. "Continuous time-series data").
                        try:
                            ds_node["attributes"] = self.attr_dict.get_attributes_for_pool_type(pool_en)
                        except Exception:
                            ds_node["attributes"] = {}
                    pool_node["datasets"].append(ds_node)

                cat_node["pools"].append(pool_node)

            out["data_chain"].append(cat_node)

        return out


# Backwards-compatible alias used by older example scripts.
DictionaryManager = GenBFKitDictManager